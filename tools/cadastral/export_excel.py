import os
from qgis.PyQt.QtWidgets import QMessageBox, QFileDialog
from qgis.PyQt.QtCore import QDateTime


class ExportExcel:
    def __init__(self, iface):
        self.iface = iface

    def run(self):
        try:
            import openpyxl
        except ImportError:
            QMessageBox.critical(None, 'Ekspor ke Excel',
                                 'openpyxl belum terinstal.\n'
                                 'Instal via: pip install openpyxl\n'
                                 'atau OSGeo4W Shell: pip install openpyxl')
            return

        layer = self.iface.activeLayer()
        if not layer:
            QMessageBox.warning(
                None,
                'Ekspor ke Excel',
                'Tidak ada layer yang aktif!')
            return

        selected = layer.selectedFeatures()
        if not selected:
            QMessageBox.warning(
                None,
                'Ekspor ke Excel',
                'Tidak ada bidang yang dipilih! Pilih bidang yang ingin diekspor.')
            return

        # Ask for save path
        default_name = f'{
            layer.name()}_{
            QDateTime.currentDateTime().toString("yyyyMMdd_HHmm")}.xlsx'
        save_path, _ = QFileDialog.getSaveFileName(
            None, 'Simpan File Excel', default_name, 'Excel Files (*.xlsx)'
        )
        if not save_path:
            return

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = layer.name()[:31]  # Excel sheet name max 31 chars

        fields = layer.fields()
        header_fill = PatternFill(
            start_color='1F4E79',
            end_color='1F4E79',
            fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        # Header row
        headers = [f.name() for f in fields]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row_idx, feature in enumerate(selected, start=2):
            for col_idx, field in enumerate(fields, start=1):
                val = feature[field.name()]
                ws.cell(row=row_idx, column=col_idx, value=val)

        # Auto column width
        for col in ws.columns:
            max_len = max((len(str(cell.value))
                          for cell in col if cell.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(
                max_len + 4, 40)

        wb.save(save_path)

        QMessageBox.information(
            None, 'Ekspor ke Excel', f'Selesai! Mengekspor {
                len(selected)} bidang ke:\n{save_path}')
